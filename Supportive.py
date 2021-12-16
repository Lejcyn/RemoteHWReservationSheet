import datetime
import pandas as pd
import os
import xlsxwriter # CANNOT BE USED FOR MODIFICATION
import openpyxl
import csv,re
from openpyxl.utils import get_column_letter

class Device:

    def __init__(self,name,mac):
        self.IPadress="NO IP"
        self.Name=name
        self.MAC=mac

    def getIP(self):
        return self.IPadress

    def getName(self):
        return self.Name

    def getMAC(self):
        return self.MAC

    def printout(self):
        print(self.getName(),self.getMAC(),self.getIP())


def CalculateExcel(number):
    return str(get_column_letter(number))

def GetWrokingWeeks(Year):#Change date if needed
    FinalList=[]
    def workdays(d, end, excluded=(6, 7)):
        days = []
        while d.date() <= end.date():
            if d.isoweekday() not in excluded:
                days.append(d)
            d += datetime.timedelta(days=1)
        return days
    TotalDays=workdays(datetime.datetime(Year, 1, 2),datetime.datetime(Year, 12, 31))
    LastElem=TotalDays[-1].isocalendar()[1]
    TotalWeeks=[[]for i in range(LastElem)]
    for data in TotalDays:
        weeknum=data.isocalendar()[1]
        TotalWeeks[weeknum-1].append(data) # Tu co nie gra

    for week in TotalWeeks:
        FinalList.append(f"{week[0].day}.{week[0].month}-{week[-1].day}.{week[0].month}")
    return FinalList

def CreateExcel(path,WorkingWeeks):
    df = pd.read_excel(r"dfTemplate.xlsx")
    writer=pd.ExcelWriter(path,engine='xlsxwriter')
    pd.set_option("max_colwidth", 40)
    workbook  = writer.book
    merge_format = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'})
    for week in WorkingWeeks:
        df.to_excel(writer,index=False,startcol=0,startrow = 1, sheet_name=week)    
        df.to_excel(writer,index=False,startcol=7,startrow = 1, sheet_name=week)
        df.to_excel(writer,index=False,startcol=14,startrow = 1, sheet_name=week)
        df.to_excel(writer,index=False,startcol=21,startrow = 1, sheet_name=week)
        df.to_excel(writer,index=False,startcol=28,startrow = 1, sheet_name=week)
        worksheet = writer.sheets[week]
        worksheet.merge_range("A1:F1","Device1", merge_format)
        worksheet.merge_range("H1:M1","Device2", merge_format)
        worksheet.merge_range("O1:T1","Device3", merge_format)
        worksheet.merge_range("V1:AA1","Device4", merge_format)
        worksheet.merge_range("AC1:AH1","Device5", merge_format)
    writer.close()

def UpdateIPs(path,WorkingWeeks,IPList):
    
    xfile = openpyxl.load_workbook(path)

    for week in WorkingWeeks:
        sheet = xfile.get_sheet_by_name(week)
        col=1
        for target in IPList:
            adress = CalculateExcel(col)+'1'
            sheet[adress]=target[0]+" "+target[1]
            col=col+7
    xfile.save(path)
    xfile.close 

def InitDevices (Devices):
    dlist=[]
    for device in Devices:
        dlist.append(Device(device[0],device[1]))
    return(dlist)

def ReadIPList(FileList):
    totalList=[]
    for file in FileList:
        with open(file) as f:
            reader = csv.reader(f)
            totalList=totalList + list(reader)  
    return totalList

def GetCurrentIPs(IPlistTXT):
    IPList=["PXIe TOP","10.92.6.29"],["PXIe Middle","10.92.6.55"],["PXIe Bottom","110.92.6.29"],["cRIO 9037-LIB","10.92.6.29"],["cRIO 9037-TSE","NO IP"]
    NewIPList= ReadIPList(IPlistTXT)
    for idx,dev in enumerate(IPList):
        for i,newdev in enumerate(NewIPList):
            if dev[0]in newdev[0]:
                dev[1]= newdev[2]
    return IPList

def InitDevices (Devices):
    dlist=[]
    for device in Devices:
        dlist.append(Device(device[0],device[1]))
    return(dlist)

def PingTargets(start,end):
    for ip in range(start,end):
        res=os.popen(f"ping 10.92.1.{ip} -w 1000 -n 1")# For office
        res=os.popen(f"ping 10.92.6.{ip} -w 1000 -n 1")#For AE Lib

def ARP(ListOfDevices):
    #Manage ARP
    f= os.popen('arp -a') 
    data = f.read()
    kek=re.findall('([-.0-9]+)\s+([-0-9a-f]{17})\s+(\w+)',data)
    for line in kek:
        for device in ListOfDevices:
            if line[1].lower() == device.getMAC().lower():
                device.IPadress=line[0]
    return ListOfDevices

def WriteIPtoFile(LogName,ListOfDevices):
    #Open File
    open(LogName, 'w+').close()
    Tfile = open(LogName, "r+")
    Tfile.truncate(0) # need '0' when using r+
    
    ARP(ListOfDevices)
    #Write to file only found devices
    for dev in ListOfDevices:
            str=dev.getIP()
            if str  != 'NO IP':
                dev.printout()
                check=dev.getName()+","+dev.getMAC()+","+dev.getIP()+"\n"
                Tfile.writelines(check)
    Tfile.close()